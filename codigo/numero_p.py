#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Oct 25 22:28:51 2022

@author: oziel
"""
from openpyxl import Workbook

wb = Workbook()
ws = wb.active
ws.title = "numeros_primos"
#print(arquivo_excel.sheetnames)
ws['A1'] = "Numeros_primos"

ite = int(input("Insira a quantidade de iterações que o código deve fazer: "))
l = 1
a = [2,3,5,7,11,13,17,19]
o = 4
while l <= ite: 
    print("iteracao: ", l)
    b = []
    c = []
    m = len(a)
    print("Tamanho de a: ",m)
    for i in range(o,(2*a[len(a) - 1])):
        print("Numero testado",i)
        k = 0
        for j in range(0,len(a)):
            d = i % a[j]
            if d == 0:
                True
            else:
                k += 1
        if k == m:
            b.append(i)
        else:
            c.append(i)
    print("*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*")
    print("Tamanho do conjunto dos numeros primos", len(a))
    print("*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*")
    print("Numeros primos a serem adicionados")
    print(b)
    print("Tamanho de b: ",len(b))
    print("*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*")
    print("Numeros compostos")
    print(c)
    
    for n in range(0, len(b)):
        a.append(b[n])
    l += 1
    o = b[len(b) - 1]
print("*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*")
print("Tamanho do conjunto dos numeros primos", len(a))
print("*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*")
print(a)
print("*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*-*")
print("Convertendo para planilha")
p = 2
for i in range(0,len(a)):
    p = str(p)
    ws['A' + p] = a[i]
    wb.save("Numeros_primos.xlsx")
    p = int(p)
    p += 1